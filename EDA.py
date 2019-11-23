import nltk
import pandas as pd
from openpyxl import load_workbook
import re
from wordcloud import WordCloud
from PIL import Image
import matplotlib.pyplot as plt
from nltk.corpus import stopwords
from nltk.corpus import webtext
from nltk.text import Text
from nltk.tokenize import TweetTokenizer
from nltk.sentiment import util
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from nltk.stem import PorterStemmer
from nltk.stem import LancasterStemmer
from nltk.stem.wordnet import WordNetLemmatizer
import pickle

# nltk.download()

def create_dataframe(tweets, pos_sent, neg_sent, neu_sent, compund_sent):
	"""
		This function creates a data frame
	"""

	data = {
		'Tweets': tweets,
		'Positive_Sentiment' : pos_sent,
		'Negative_Sentiment' : neg_sent, 
		'Neutral_Sentiment'	: neu_sent, 
		'Compound_Sentiment': compund_sent
	}

	tweets_df = pd.DataFrame(data)

	return tweets_df

def get_tweets(file_name):
	"""
		This function return the list of tweets
	"""
	wb = load_workbook(file_name)
	sheet = wb['Sheet']

	tweets = []

	rows = sheet.max_row

	for r in range(rows):
		if r > 1:
			c = sheet.cell(r, 1)
			tweets.append(c.value)

	return tweets


def clean_tweets(tweets):
	"""
		data preprocessing
		This function return the list of clean tweets - removed duplicate tweets
	"""
	sw = set(stopwords.words('english'))
	tknzr = TweetTokenizer()
	lmtzr = WordNetLemmatizer()
	# print("Length of tweets: " + str(len(tweets)))

	cleaned_tweets = []

	for tweet in tweets:
		# text = Text(tweet)

		# remove all charactes except words
		tweet = re.sub('[^a-zA-Z]', " ", tweet)

		# stripe all extra spaces
		tweet = tweet.strip()

		# lemmatize the word, change to lower case and remove all stop words and words of length 1
		words = [lmtzr.lemmatize(w.lower()) for w in tknzr.tokenize(tweet) if w.lower() not in sw and len(w) > 1]

		formatted_text = " "
		formatted_text = formatted_text.join(words)

		# print(formatted_text)
		# print("===================")
		
		cleaned_tweets.append(formatted_text)

		"""
			Removing duplicate tweets
		"""

	return list(set(cleaned_tweets))

def get_sentiments(tweets):
	sid_obj = SentimentIntensityAnalyzer()
	sentiment_pos = []
	sentiment_neg = []
	sentiment_neu = []
	sentiment_compound = []


	for tweet in tweets:
		
		# print(tweet)
		sentiment_dict = sid_obj.polarity_scores(tweet)
		sentiment_pos.append(sentiment_dict['pos'])
		sentiment_neg.append(sentiment_dict['neg'])
		sentiment_neu.append(sentiment_dict['neu'])
		sentiment_compound.append(sentiment_dict['compound'])


	return sentiment_pos, sentiment_neg, sentiment_neu, sentiment_compound


def classify_sentiments(tweets_df):
	is_positive = []
	is_negative = []
	is_neutral = []

	for i in tweets_df.index:
		sentiment_list = []
		sentiment_list.append(tweets_df['Positive_Sentiment'][i])
		sentiment_list.append(tweets_df['Negative_Sentiment'][i])
		sentiment_list.append(tweets_df['Neutral_Sentiment'][i])

		ind = sentiment_list.index(max(sentiment_list))
		
		if ind == 0:
			is_positive.append(1)
			is_negative.append(0)
			is_neutral.append(0)

		elif ind == 1:
			is_positive.append(0)
			is_negative.append(1)
			is_neutral.append(0)
		else:
			is_positive.append(0)
			is_negative.append(0)
			is_neutral.append(1)


	return is_positive, is_negative, is_neutral	

		

def create_wordcloud(tweets):
	text = ''

	for tweet in tweets:
		text = text + tweet + " "

	# Plot wordcloud 
	wordcloud = WordCloud(width=400, height=400,max_font_size=50, max_words=100, 
	                      background_color="white").generate(text)

	plt.figure()
	plt.imshow(wordcloud, interpolation='bilinear')
	plt.axis('off')
	plt.show()


# def to_picke(tweets_df, file_name):
# 	"""
# 		Function to save the dataframe to a pickle_file
# 	"""
# 	with open(file_name, 'wb') as pickle_file:
# 		pickle.dump(tweets_df, pickle_file)
# 		pickle_file.close()


if __name__ == '__main__':
	
	tweets = get_tweets('tweets.xlsx')	# Get list of tweets present in file tweets.xlsx

	tweets_clean = clean_tweets(tweets)	# Get list of clean tweets

	print("type: " + str(type(tweets_clean)))

	print("Total number of tweets after cleaning... " + str(len(tweets)))

	sentiment_pos, sentiment_neg, sentiment_neu, sentiment_compound = get_sentiments(tweets_clean)

	tweets_df = create_dataframe(tweets_clean, sentiment_pos, sentiment_neg, sentiment_neu, sentiment_compound)

	print("After DataFrame is created")
	print("Size of data frame: " + str(tweets_df.size))
	print("Shape of data frame: " + str(tweets_df.shape))
	print("==================================")

	is_positive, is_negative, is_neutral = classify_sentiments(tweets_df)

	tweets_df['is_positive'] = is_positive
	tweets_df['is_negative'] = is_negative
	tweets_df['is_neutral'] = is_neutral

	print("After classify_sentiments is created")
	print("Size of data frame: " + str(tweets_df.size))
	print("Shape of data frame: " + str(tweets_df.shape))
	print("==================================")

	print(tweets_df.iloc[:5, 1: ])
	# # create_wordcloud(tweets)	# This function will create a word cloud

	pd.to_pickle(tweets_df, './tweets_df.pkl')

