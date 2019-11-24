import tweepy
import twitter_cred
from textblob import TextBlob
import openpyxl
from openpyxl import load_workbook
import preprocessor as p
import re
import nltk
import datetime
from datetime import timedelta, date
from datetime import datetime
import pandas as pd
import pickle
import json


def auth_twitter():
    """
    Function for authenticating twitter account
    :return:
    """
    auth = tweepy.OAuthHandler(twitter_cred.CON_API_KEY, twitter_cred.CON_API_KEY_SECRET)
    auth.set_access_token(twitter_cred.ACCESS_TOKEN, twitter_cred.ACCESS_TOKEN_SECRET)
    return auth


def get_api(auth):
    """
    Fuction that returns the twitter api
    :param auth:
    :return:
    """
    api = tweepy.API(auth)
    return api


def get_public_tweets(api, search_query, tweet_count):
    """
    Fuction returns the public tweets
    :param api:
    :param search_query:
    :return:
    """
    # public_tweets = api.search(search_query, count=1000, lang='en')

    public_tweets = tweepy.Cursor(api.search, q=search_query, lang='en').items(tweet_count)

    return public_tweets


def get_tweet_count_date(api, search_query, max_tweet_count, since, days):
	count_per_day = []
	date_list = []

	for single_date in (since + timedelta(n) for n in range(days)):
		date_list.append(datetime.strftime(single_date.date(), '%b %d'))
		count = 0
		next_date = single_date + timedelta(1)
		tweets = tweepy.Cursor(api.search, q=search_query, lang='en', since=single_date, until=next_date).items(max_tweet_count)

		for tweet in tweets:
			count = count + 1

		count_per_day.append(count)
			

	return count_per_day, date_list


def display_tweets(public_tweets):
    """
    Fuction displays the first 100 tweets
    :param public_tweets:
    :return:
    """
    count = 1
    for tweet in public_tweets:
        if count < 10:
            print(tweet.created_at)
            print("=====================")
        else:
            break

        count = count + 1


def get_data(tweets):
    """
    Function returns the tweet, its polarity and its subjectivity

    :param tweets:
    :return:
    """
    jnu_tweets = []
    tweets_polarity = []
    tweets_subjectivity = []

    count = 1

    for tweet in tweets:
    	t = p.clean(tweet.text)

    	# removing colons from the tweet
    	t = re.sub(r':', '', t)

    	# removing mentions
    	t = re.sub(r'‚Ä¶', '', t)

    	# removing non ascii characters
    	t = re.sub(r'[^\x00-\x7F]+',' ', t)

    	jnu_tweets.append(t)
    	tweets_polarity.append(TextBlob(t).polarity)
    	tweets_subjectivity.append(TextBlob(t).subjectivity)
        

    return jnu_tweets, tweets_polarity, tweets_subjectivity


def create_file(file_name, public_tweets):
    """
    This function simply creates a excel file having three columns of tweet, polarity and subjectivity

    :param file_name:
    :param public_tweets:
    :return:
    """
    tweets, polarity, subjectivity = get_data(public_tweets)
    wb = openpyxl.Workbook()
    sheet = wb.active

    c = sheet.cell(1, 1)
    c.value = "Tweets"

    c = sheet.cell(1, 2)
    c.value = "Polarity"

    c = sheet.cell(1, 3)
    c.value = "Subjectivity"

    row = 2
    for t in tweets:
        c = sheet.cell(row, 1)
        c.value = t
        row = row + 1

    row = 2
    for p in polarity:
        c = sheet.cell(row, 2)
        c.value = p
        row = row + 1

    row = 2
    for s in subjectivity:
        c = sheet.cell(row, 3)
        c.value = s
        row = row + 1

    wb.save(file_name)
    print("File created successfully")


if __name__ == '__main__':

	search_query = '#standwithjnu OR #JNUFeeHike OR #JNUProtests OR #Jnu'


	print("Authenticating ....")
	twitter_auth = auth_twitter()
	print("Authenticated.")


	twitter_api = get_api(twitter_auth)

	print("Fetching tweets....")

	"""
		Fetching 2000 tweets
	"""
	max_tweets = 2000
	twitter_tweets = get_public_tweets(twitter_api, search_query, max_tweets)
	print("Fetched.")

	# print("Tweets locations")
	# dissplay_tweets(twitter_tweets)


	print("Creating xlsx sheet.....")
	create_file("tweets.xlsx", twitter_tweets)
	print("File created and data saved successfully.")

	since = datetime(2019, 11, 17, 0, 0, 0)
	until = datetime(2019, 11, 24, 0, 0, 0)

	days = 7

	print("Counting tweets of each day of week 17th - 24th Nov")
	count_per_day, date_list = get_tweet_count_date(twitter_api, 'jnu', 1000, since, days)
	print("Tweets counted.")
	
	data_df = {
		'Date' : date_list, 
		'Tweet_Count': count_per_day
	}
	
	count_per_day_df = pd.DataFrame(data_df)

	print("Saving data to count_per_day_df.pkl file")
	pd.to_pickle(count_per_day_df, './count_per_day_df.pkl')
	print("Data Saved.")
